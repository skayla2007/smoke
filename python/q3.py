"""
问题3：FY1 投放 3 枚烟幕干扰弹，对 M1 实施干扰。

跟问题2的本质区别（硬约束）：
同一架无人机全程只能选一个 (speed, heading)，3 枚弹的起爆点因此被钉死在同一条
射线上——不能像问题2那样给每枚弹独立配一条射线、各自反解最优起爆点。
bomb_detonation 的公式说明了这一点：给定 (speed,heading) 后，
  det_xy 只由 det_time = t_drop+t_fuze 决定（沿射线走多远）
  det_z  只由 t_fuze 决定（自由落体多久）
所以 3 枚弹真正能自由选的是各自的 (t_drop_i, t_fuze_i)，heading/speed 是公共的。

关键发现（写代码前先用交互式脚本验证过，见对话记录）：
1. 问题2里"单发最优"的那两条射线（heading≈5°、≈177°），沿途的有效窗口只有约
   1.7~2s 宽——塞不下 3 发间隔≥1s 还都高质量的弹。
2. 对 (heading,speed) 做一次粗筛扫描（用 batch_quick_score 量化"沿途窗口宽度"）
   发现 heading=180° 附近窗口最宽（~5~7s），原因是这个方向恰好和导弹的地面
   航迹方向平行（导弹从 (20000,0,2000) 直指假目标原点，水平投影方向正好是
   180°）——无人机与导弹"同向而行"，相对接近速度慢很多，几何关系变化得慢，
   窗口自然宽。这不是巧合数字，而是"平行于目标航迹方向"这个可解析算出的方向。
3. 但即便窗口宽，3发弹如果贪心地"各自选最优 t_fuze、只留 1s 最小间隔"，
   窗口会互相大量重叠，并集反而更差；需要故意加大间隔（超过最小值）来错开
   窗口，用并集长度作为目标去优化，而不是让每发弹独立最优。

做法（两层）：
  1. 候选"射线"生成：复用问题2的 geometric_candidates+cluster_candidates
     （得到几条"至少单发能用"的方向），再加上解析算出的"平行于导弹航迹"方向
     作为额外候选（这条不是从数据里蒙出来的，是从导弹初始位置和假目标位置
     直接算出来的，具有一般性）。
  2. 对每条候选射线：粗筛出"起爆时刻→最佳t_fuze"剖面，网格搜索 (t_drop1,
     gap2, gap3) 的组合（gap>=MIN_BOMB_INTERVAL，但不限制上限，允许远超最小
     间隔），用精确的 union_length 评估，取最优组合做初始解；再对
     (speed, heading, t_drop1, gap2, gap3, t_fuze1, t_fuze2, t_fuze3) 8 个
     变量做带边界的 Nelder-Mead 联合精修（gap 重参数化，把"间隔>=1s"这个
     约束变成普通的下界，不需要额外写约束优化）。
  3. 比较所有候选射线精修后的结果，取全局最优，写入 result1.xlsx。
"""

import numpy as np
import openpyxl
from scipy.optimize import minimize

from model import (
    F_single_bomb,
    bomb_detonation,
    shielding_intervals,
    union_length,
    batch_quick_score,
    UAV_SPEED_MIN,
    UAV_SPEED_MAX,
    MIN_BOMB_INTERVAL,
    MISSILE_INIT,
    FAKE_TARGET,
)
from q2 import geometric_candidates, cluster_candidates, objective as q2_objective

UAV_NAME = "FY1"
MISSILE_NAME = "M1"
N_BOMBS = 3

# 决策变量 X = (speed, heading, t_drop1, gap2, gap3, t_fuze1, t_fuze2, t_fuze3)
# t_drop2 = t_drop1+gap2, t_drop3 = t_drop2+gap3；gap>=MIN_BOMB_INTERVAL 保证间隔约束。
BOUNDS = [
    (UAV_SPEED_MIN, UAV_SPEED_MAX),
    (0.0, 2 * np.pi),
    (0.0, 30.0),
    (MIN_BOMB_INTERVAL, 30.0),
    (MIN_BOMB_INTERVAL, 30.0),
    (0.001, 20.0),
    (0.001, 20.0),
    (0.001, 20.0),
]


def parallel_to_missile_heading(missile_name: str) -> float:
    """
    解析算出"与导弹地面航迹平行"的方向角：导弹从初始点直飞向假目标，
    这条方向的水平投影就是候选射线之一（无人机与导弹同向而行，相对接近速度慢，
    沿途窗口更宽，适合塞多发弹）。
    """
    p0 = MISSILE_INIT[missile_name]
    dxy = FAKE_TARGET[:2] - p0[:2]
    return float(np.arctan2(dxy[1], dxy[0]))


def union_objective(x: np.ndarray, n_coarse: int = 300) -> float:
    """负的3发弹并集遮蔽时长（供最小化算法使用）。"""
    speed, heading, t_drop1, gap2, gap3, tf1, tf2, tf3 = x
    t_drop2 = t_drop1 + gap2
    t_drop3 = t_drop2 + gap3
    groups = []
    for td, tf in [(t_drop1, tf1), (t_drop2, tf2), (t_drop3, tf3)]:
        _, det_point, det_time = bomb_detonation(UAV_NAME, speed, heading, td, tf)
        if det_point[2] < 0:
            continue
        groups.append(
            shielding_intervals(MISSILE_NAME, det_point, det_time, n_coarse=n_coarse)
        )
    return -union_length(groups)


def ray_profile(
    speed: float, heading: float, td_grid: np.ndarray, tf_grid: np.ndarray
) -> np.ndarray:
    """
    向量化粗筛：固定 (speed,heading)，对 td_grid 中每个 t_drop 找 tf_grid 里最好的 t_fuze。
    返回形状 (len(td_grid),) 的最佳 t_fuze 数组（供后续初始解组装用）。
    """
    TD, TF = np.meshgrid(td_grid, tf_grid, indexing="ij")
    scores = batch_quick_score(
        UAV_NAME,
        MISSILE_NAME,
        np.full(TD.size, speed),
        np.full(TD.size, heading),
        TD.ravel(),
        TF.ravel(),
        n_t=30,
    ).reshape(TD.shape)
    best_tf_idx = scores.argmax(axis=1)
    return tf_grid[best_tf_idx]


def best_initial_guess_for_ray(
    speed: float, heading: float
) -> tuple[np.ndarray, float] | None:
    """
    给定一条候选射线 (speed, heading)，搜索 (t_drop1, gap2, gap3) 组合（配合剖面估计
    的 t_fuze），用精确的 union_length 评估，返回最优的初始 8 维解及其并集时长。
    找不到任何非零方案时返回 None。
    """
    td_grid = np.linspace(0.0, 20.0, 80)
    tf_grid = np.linspace(0.02, 10.0, 100)
    best_tf_of_td = ray_profile(speed, heading, td_grid, tf_grid)

    def approx_tf(td: float) -> float:
        return float(np.interp(td, td_grid, best_tf_of_td))

    best: tuple[np.ndarray, float] | None = None
    t_drop1_options = np.linspace(0.0, 8.0, 5)
    gap_options = np.array([1.0, 2.0, 3.0, 4.0, 5.0, 6.0])
    for t_drop1 in t_drop1_options:
        for gap2 in gap_options:
            for gap3 in gap_options:
                t_drop2 = t_drop1 + gap2
                t_drop3 = t_drop2 + gap3
                tfs = [approx_tf(t_drop1), approx_tf(t_drop2), approx_tf(t_drop3)]
                groups = []
                for td, tf in zip([t_drop1, t_drop2, t_drop3], tfs):
                    _, det_point, det_time = bomb_detonation(
                        UAV_NAME, speed, heading, td, tf
                    )
                    if det_point[2] < 0:
                        continue
                    groups.append(
                        shielding_intervals(
                            MISSILE_NAME, det_point, det_time, n_coarse=300
                        )
                    )
                u = union_length(groups)
                if u > 0 and (best is None or u > best[1]):
                    x0 = np.array(
                        [speed, heading, t_drop1, gap2, gap3, tfs[0], tfs[1], tfs[2]]
                    )
                    best = (x0, u)
    return best


def polish(x0: np.ndarray, n_coarse: int = 200) -> tuple[np.ndarray, float]:
    """
    n_coarse=200 精修实测和 800 几乎同样精确（差异<0.001s），但快 3.5 倍；
    maxiter=3500 足以让8维 Nelder-Mead 实际收敛（示例跑到 nit≈2881 就 success=True）。
    """
    result = minimize(
        union_objective,
        x0,
        args=(n_coarse,),
        method="Nelder-Mead",
        bounds=BOUNDS,
        options={"xatol": 1e-5, "fatol": 1e-7, "maxiter": 3500},
    )
    return result.x, -result.fun


def collect_candidate_rays(rng: np.random.Generator) -> list[tuple[float, float]]:
    """
    收集候选 (speed, heading) 射线：
    1. 复用问题2的几何候选生成+聚类，得到"至少单发能用"的方向；
    2. 加上解析算出的"与导弹航迹平行"方向（配几个不同速度），
       这条不依赖数据，是从物理机制直接推出来的候选。
    """
    feasible = geometric_candidates(rng=rng, n_samples=100_000)
    scores = np.array([-q2_objective(x, n_coarse=60) for x in feasible])
    nonzero = feasible[scores > 0]
    rays = []
    if len(nonzero) > 0:
        clusters = cluster_candidates(nonzero)
        for cluster in clusters:
            cluster_scores = np.array([-q2_objective(x, n_coarse=150) for x in cluster])
            rep = cluster[np.argmax(cluster_scores)]
            rays.append((float(rep[0]), float(rep[1])))

    parallel_heading = parallel_to_missile_heading(MISSILE_NAME)
    for speed in (UAV_SPEED_MIN, 100.0, 130.0, UAV_SPEED_MAX):
        rays.append((speed, parallel_heading))
    return rays


if __name__ == "__main__":
    rng = np.random.default_rng(42)

    rays = collect_candidate_rays(rng)
    print("候选射线数: %d" % len(rays), flush=True)

    initial_guesses = []
    for speed, heading in rays:
        result = best_initial_guess_for_ray(speed, heading)
        if result is not None:
            x0, u = result
            initial_guesses.append((x0, u))
            print(
                "  射线 speed=%.1f heading=%.4f(%.1fdeg) -> 粗解并集=%.4f"
                % (speed, heading, np.degrees(heading) % 360, u),
                flush=True,
            )

    polished = []
    for i, (x0, _) in enumerate(initial_guesses):
        x, u = polish(x0)
        polished.append((x, u))
        print(
            "  已精修 %d/%d，并集=%.4f" % (i + 1, len(initial_guesses), u), flush=True
        )
    polished.sort(key=lambda item: item[1], reverse=True)

    print("\n各候选射线精修后的并集时长（按从优到劣排序，前5个）：")
    for x, u in polished[:5]:
        speed, heading, t_drop1, gap2, gap3, tf1, tf2, tf3 = x
        print(
            "  union=%.4f speed=%.2f heading=%.2fdeg t_drop=(%.3f,%.3f,%.3f) t_fuze=(%.3f,%.3f,%.3f)"
            % (
                u,
                speed,
                np.degrees(heading) % 360,
                t_drop1,
                t_drop1 + gap2,
                t_drop1 + gap2 + gap3,
                tf1,
                tf2,
                tf3,
            )
        )

    best_x, best_u = polished[0]
    speed, heading, t_drop1, gap2, gap3, tf1, tf2, tf3 = best_x
    bombs = [
        (t_drop1, tf1),
        (t_drop1 + gap2, tf2),
        (t_drop1 + gap2 + gap3, tf3),
    ]

    print("\n===== 问题3 最优投放策略 =====")
    print("飞行方向: %.4f rad (%.2f deg)" % (heading, np.degrees(heading) % 360))
    print("飞行速度: %.4f m/s" % speed)
    rows = []
    for i, (td, tf) in enumerate(bombs, start=1):
        duration, drop_point, det_point, det_time = F_single_bomb(
            UAV_NAME, MISSILE_NAME, speed, heading, td, tf, n_coarse=4000
        )
        print(
            "  弹%d: t_drop=%.4f t_fuze=%.4f 投放点=%s 起爆点=%s 单发遮蔽=%.4f"
            % (i, td, tf, drop_point, det_point, duration)
        )
        rows.append((drop_point, det_point, duration))
    print("3发并集总遮蔽时长: %.4f s" % best_u)

    # ---- 写入 result1.xlsx ----
    wb = openpyxl.load_workbook("result1.xlsx")
    ws = wb.active
    heading_deg = np.degrees(heading) % 360
    for i, (drop_point, det_point, duration) in enumerate(rows):
        r = i + 2  # 表头占第1行，弹1/2/3对应第2~4行
        ws.cell(row=r, column=1, value=float(heading_deg))
        ws.cell(row=r, column=2, value=float(speed))
        ws.cell(row=r, column=3, value=i + 1)
        ws.cell(row=r, column=4, value=float(drop_point[0]))
        ws.cell(row=r, column=5, value=float(drop_point[1]))
        ws.cell(row=r, column=6, value=float(drop_point[2]))
        ws.cell(row=r, column=7, value=float(det_point[0]))
        ws.cell(row=r, column=8, value=float(det_point[1]))
        ws.cell(row=r, column=9, value=float(det_point[2]))
        ws.cell(row=r, column=10, value=float(duration))
    wb.save("result1.xlsx")
    print("\n结果已写入 result1.xlsx")
