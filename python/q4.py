"""
问题4：FY1、FY2、FY3 各投放 1 枚烟幕干扰弹，对 M1 实施干扰。

跟问题3不同，这里没有"同机多弹共享射线"的硬约束——3架无人机各自独立选择
方向、速度、投放/引信延迟，互不影响。所以可以直接复用问题2的整套逻辑
（geometric_candidates 反解候选起爆配置 -> cluster_candidates 分孤岛 ->
逐个 Nelder-Mead 精修），对每架无人机分别跑一遍即可，取值时把 UAV_NAME
从硬编码的 "FY1" 换成了函数参数（见 q2.py 的 solve_single_uav_single_bomb）。

3架机各自的最优解互相独立，最终答案就是各自跑出来的最优解拼在一起，
不需要额外的联合优化——这也是问题4比问题3简单的地方。
"""

from pathlib import Path

import numpy as np
import openpyxl

from model import F_single_bomb, shielding_intervals, union_length
from q2 import solve_single_uav_single_bomb

UAVS = ["FY1", "FY2", "FY3"]
MISSILE_NAME = "M1"

# 项目结构是 <repo根>/python/q4.py，而 result2.xlsx 在 <repo根> 下，
# 用脚本自身路径定位，不依赖运行时的当前工作目录。
RESULT_PATH = Path(__file__).resolve().parent.parent / "result2.xlsx"

Row = tuple[float, float, np.ndarray, np.ndarray, float]

if __name__ == "__main__":
    rng = np.random.default_rng(42)

    rows: list[Row | None] = []
    interval_groups = []
    for uav_name in UAVS:
        polished = solve_single_uav_single_bomb(uav_name, MISSILE_NAME, rng)
        if not polished:
            print("%s: 未找到非零遮蔽解" % uav_name)
            rows.append(None)
            continue

        best_x, _ = polished[0]
        speed, heading, t_drop, t_fuze = best_x
        duration, drop_point, det_point, det_time = F_single_bomb(
            uav_name, MISSILE_NAME, speed, heading, t_drop, t_fuze, n_coarse=8000
        )
        heading_deg = np.degrees(heading) % 360

        print(
            "%s: 方向=%.2fdeg 速度=%.2fm/s t_drop=%.4f t_fuze=%.4f "
            "投放点=%s 起爆点=%s 遮蔽=%.4fs"
            % (
                uav_name,
                heading_deg,
                speed,
                t_drop,
                t_fuze,
                drop_point,
                det_point,
                duration,
            )
        )
        rows.append((heading_deg, speed, drop_point, det_point, duration))
        interval_groups.append(
            shielding_intervals(MISSILE_NAME, det_point, det_time, n_coarse=8000)
        )

    # 3 架机各自独立选的最优解，起爆时刻不一定错开——用精确区间求并集，
    # 而不是想当然地把各自时长相加（若窗口重叠，简单相加会高估）。
    total_union = union_length(interval_groups)
    naive_sum = sum(r[-1] for r in rows if r is not None)
    print(
        "\n三机独立最优解的并集总遮蔽时长: %.4f s（各自时长直接相加是 %.4f s，"
        "差值 %.4f s 就是窗口重叠部分）"
        % (total_union, naive_sum, naive_sum - total_union)
    )

    # ---- 写入 result2.xlsx ----
    wb = openpyxl.load_workbook(RESULT_PATH)
    ws = wb.active
    for i, row in enumerate(rows):
        if row is None:
            continue
        heading_deg, speed, drop_point, det_point, duration = row
        r = i + 2  # 表头占第1行，FY1/FY2/FY3 对应第2~4行
        ws.cell(row=r, column=2, value=float(heading_deg))
        ws.cell(row=r, column=3, value=float(speed))
        ws.cell(row=r, column=4, value=float(drop_point[0]))
        ws.cell(row=r, column=5, value=float(drop_point[1]))
        ws.cell(row=r, column=6, value=float(drop_point[2]))
        ws.cell(row=r, column=7, value=float(det_point[0]))
        ws.cell(row=r, column=8, value=float(det_point[1]))
        ws.cell(row=r, column=9, value=float(det_point[2]))
        ws.cell(row=r, column=10, value=float(duration))
    wb.save(RESULT_PATH)
    print("\n结果已写入 result2.xlsx")
